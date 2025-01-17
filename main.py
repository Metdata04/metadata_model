from flask import Flask, request, render_template, redirect, send_from_directory, url_for
import pandas as pd
from openpyxl import load_workbook, Workbook
import os
import requests
import base64

app = Flask(__name__)

# Directories to store files
METADATA_FOLDER = "/tmp/metadata"
METADATA_REPORTS_FOLDER = "/tmp/metadata_reports"

os.makedirs(METADATA_FOLDER, exist_ok=True)
os.makedirs(METADATA_REPORTS_FOLDER, exist_ok=True)

# GitHub repository details
GITHUB_API_URL = "https://api.github.com"
REPO_OWNER = "Metdata04"  # GitHub username
REPO_NAME = "metadata_model"  # Repository name
BRANCH_NAME = "main"  # Branch name
GITHUB_TOKEN = os.getenv("Metadata_token")  # GitHub personal access token
if not GITHUB_TOKEN:
    raise ValueError("GitHub token is not set in environment variables.")

def upload_file_to_github(file_path, file_name, commit_message="Add new report"):
    url = f"{GITHUB_API_URL}/repos/{REPO_OWNER}/{REPO_NAME}/contents/{file_name}"

    # Check if the file exists
    response = requests.get(
        url,
        headers={"Authorization": f"token {GITHUB_TOKEN}"}
    )

    sha = None  # Default to None if file doesn't exist
    if response.status_code == 200:
        # File exists, get the sha of the current file
        sha = response.json().get('sha')

    # Encode the file to base64
    with open(file_path, "rb") as file:
        encoded_file = base64.b64encode(file.read()).decode("utf-8")

    # Prepare the data for upload (either creating or updating the file)
    data = {
        "message": commit_message,
        "content": encoded_file,
        "branch": BRANCH_NAME
    }

    # If sha is found, it means we are updating an existing file
    if sha:
        data['sha'] = sha

    # Perform the PUT request to upload or update the file
    response = requests.put(
        url,
        headers={"Authorization": f"token {GITHUB_TOKEN}"},
        json=data
    )

    # Check the response status
    if response.status_code in (200, 201):
        return f"File {file_name} uploaded/updated successfully."
    else:
        raise Exception(f"Failed to upload file: {response.status_code} - {response.text}")

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/images/<path:filename>')
def serve_static(filename):
    print(f"Serving file: {os.path.join('public/images', filename)}")
    return send_from_directory('public/images', filename)

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return "No file part", 400

    file = request.files['file']

    if file.filename == '':
        return "No selected file", 400

    input_file_path = os.path.join(METADATA_FOLDER, file.filename)
    file.save(input_file_path)

    station_name = os.path.splitext(file.filename)[0]
    report_file_path = os.path.join(METADATA_REPORTS_FOLDER, f"{station_name}_Metadata_Report.xlsx")

    try:
        generate_availability_report(input_file_path, report_file_path, station_name)
        upload_result = upload_file_to_github(report_file_path, f"{station_name}_Metadata_Report.xlsx", f"Add/update metadata report for {station_name}")
        print(upload_result)
        return redirect(url_for('report_generated', station_name=station_name))
    except Exception as e:
        return f"Error generating the report: {e}", 500

@app.route('/report_generated/<station_name>')
def report_generated(station_name):
    return f"Report successfully generated for station: {station_name}. You can find it in the GitHub repository."

def generate_availability_report(input_file, report_file, station_name):
    # Read the input CSV file
    df = pd.read_csv(input_file)

    # Ensure the 'Date' column exists and is valid
    if 'Date' not in df.columns:
        raise ValueError("The input file must contain a 'Date' column.")

    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    if df['Date'].isna().any():
        raise ValueError("Some 'Date' values in the 'Date' column are invalid.")

    # Create "Year" and "Month" columns for grouping
    df['Year'] = df['Date'].dt.year
    df['Month'] = df['Date'].dt.strftime('%b').str.upper()

    # Check if the report file exists in GitHub
    url = f"{GITHUB_API_URL}/repos/{REPO_OWNER}/{REPO_NAME}/contents/{station_name}_Metadata_Report.xlsx"
    response = requests.get(url, headers={"Authorization": f"token {GITHUB_TOKEN}"})

    if response.status_code == 200:
        # File exists, download it
        content = base64.b64decode(response.json()['content'])
        with open(report_file, "wb") as f:
            f.write(content)
        wb = load_workbook(report_file)
        ws = wb.active
    else:
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Metadata Report"
        headers = ["Year", "Month", "Data Availability", "Data Missing", "Station Name"] + [
            "Outdoor Temperature (°C)", "Feels Like (°C)", "Dew Point (°C)", "Wind Speed (km/hr)", "Wind Gust (km/hr)",
            "Max Daily Gust (km/hr)", "Wind Direction (°)", "Rain Rate(mm/hr)", "Event Rain (mm)", "Daily Rain (mm)",
            "Weekly Rain (mm)", "Monthly Rain (mm)", "Yearly Rain (mm)", "Relative Pressure (hPa)", "Humidity (%)",
            "Ultra-Violet Radiation Index", "Solar Radiation (W/m^2)", "Indoor Temperature (°C)", "Indoor Humidity (%)",
            "PM2.5 Outdoor (µg/m³)", "PM2.5 Outdoor 24 Hour Average (µg/m³)", "Indoor Battery", "Indoor Feels Like (°C)",
            "Indoor Dew Point (°C)", "Absolute Pressure (hPa)", "Outdoor Battery", "Avg Wind Direction (10 mins) (°)",
            "Avg Wind Speed (10 mins) (km/hr)", "Total Rain", "CO2 Battery", "PM 2.5 (µg/m³)"
        ]
        ws.append(headers)


    # Group data by Year and Month
    grouped = df.groupby(['Year', 'Month'])
    for (year, month), month_data in grouped:
        # Calculate data availability and missing data
        available_dates = month_data['Date'].dt.strftime('%d/%m').tolist()
        data_availability = f"{available_dates[0]}-{available_dates[-1]}" if available_dates else "-"

        # Create the full month for comparison to get missing dates
        month_number = month_data['Date'].dt.month.iloc[0]
        all_dates = pd.date_range(
            start=f"{year}-{month_number:02d}-01",
            end=f"{year}-{month_number:02d}-{month_data['Date'].dt.days_in_month.iloc[0]}"
        )

        all_dates_set = set(all_dates.date)
        available_dates_set = set(month_data['Date'].dt.date)
        missing_dates = all_dates_set - available_dates_set
        formatted_missing_dates = [d.strftime('%d/%m') for d in sorted(missing_dates)]
        data_missing = ", ".join(formatted_missing_dates) if formatted_missing_dates else "-"

          # Filter the columns for available variables for this station
        expected_variables = [
            "Outdoor Temperature (°C)", "Feels Like (°C)", "Dew Point (°C)", "Wind Speed (km/hr)", "Wind Gust (km/hr)",
            "Max Daily Gust (km/hr)", "Wind Direction (°)", "Rain Rate(mm/hr)", "Event Rain (mm)", "Daily Rain (mm)",
            "Weekly Rain (mm)", "Monthly Rain (mm)", "Yearly Rain (mm)", "Relative Pressure (hPa)", "Humidity (%)",
            "Ultra-Violet Radiation Index", "Solar Radiation (W/m^2)", "Indoor Temperature (°C)", "Indoor Humidity (%)",
            "PM2.5 Outdoor (µg/m³)", "PM2.5 Outdoor 24 Hour Average (µg/m³)", "Indoor Battery", "Indoor Feels Like (°C)",
            "Indoor Dew Point (°C)", "Absolute Pressure (hPa)", "Outdoor Battery", "Avg Wind Direction (10 mins) (°)",
            "Avg Wind Speed (10 mins) (km/hr)", "Total Rain", "CO2 Battery", "PM 2.5 (µg/m³)"
        ]
        available_variables = [col for col in expected_variables if col in month_data.columns]

        # Append the station's data to the Excel sheet
        row = [year, month, data_availability, data_missing, station_name]
        for variable in expected_variables:
            row.append("✓" if variable in available_variables else "-")


        ws.append(row)

    # Save the updated workbook
    wb.save(report_file)


if __name__ == '__main__':
    app.run(debug=True)
